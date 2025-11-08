import openpyxl

# Archivos origen y destino
archivo_origen = r"D:\dev\Control_vacunas\BASE DE DATOS GARANTIA DE DERECHOS SEMILLAS.xlsx"
archivo_destino = r"D:\dev\Control_vacunas\DATA_FAKE.xlsx"

# Cargar ambos libros
hoja_origen_nombre = "CUNAS 1"

# Cargar ambos libros (sin evaluar fórmulas)
wb_origen = openpyxl.load_workbook(archivo_origen, data_only=False)
wb_destino = openpyxl.load_workbook(archivo_destino)

# Seleccionar la hoja origen
hoja_origen = wb_origen[hoja_origen_nombre]

# 🔹 Recorrer todas las hojas del archivo destino
for hoja_destino in wb_destino.worksheets:
    print(f"Copiando fórmulas en: {hoja_destino.title}")
    
    # Copiar fórmulas celda por celda
    for fila in hoja_origen.iter_rows():
        for celda in fila:
            if celda.data_type == "f":  # si la celda contiene una fórmula
                hoja_destino[celda.coordinate].value = celda.value

# Guardar cambios
wb_destino.save(archivo_destino)
print("✅ Fórmulas copiadas correctamente desde CUNAS1 a todas las hojas del destino.")